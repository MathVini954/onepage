# src/excel_reader.py
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.utils import norm, is_blank, to_month, to_float, fmt_brl


def load_wb(path: str | Path) -> Workbook:
    path = Path(path)
    keep_vba = path.suffix.lower() == ".xlsm"
    return load_workbook(path, data_only=True, keep_vba=keep_vba)


def sheetnames(wb: Workbook) -> list[str]:
    ignore = {"LEIA-ME", "README", "READ ME"}
    return [s for s in wb.sheetnames if norm(s) not in ignore]


def _find_row_contains(ws: Worksheet, needle: str, col: int = 1, max_row: int = 250) -> int | None:
    n = norm(needle)
    for r in range(1, min(ws.max_row, max_row) + 1):
        if n in norm(ws.cell(r, col).value):
            return r
    return None


def _find_header_row(ws: Worksheet, headers: list[str], max_row: int = 400, max_col: int = 80) -> tuple[int | None, list[str] | None]:
    want = [norm(h) for h in headers]
    r_max = min(ws.max_row, max_row)
    c_max = min(ws.max_column, max_col)

    for r in range(1, r_max + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, c_max + 1)]
        ok = True
        for h in want:
            if h and h not in row_vals:
                ok = False
                break
        if ok:
            return r, row_vals
    return None, None


def _col_idx(row_vals: list[str], header: str) -> int | None:
    h = norm(header)
    for i, v in enumerate(row_vals, start=1):
        if v == h:
            return i
    return None


def read_resumo_financeiro(ws: Worksheet) -> dict[str, float | None]:
    title_row = _find_row_contains(ws, "RESUMO FINANCEIRO", col=1, max_row=120)
    if title_row is None:
        return {}

    mapping = {
        "ORÇAMENTO INICIAL": "ORÇAMENTO INICIAL (R$)",
        "ORCAMENTO INICIAL": "ORÇAMENTO INICIAL (R$)",
        "ORÇAMENTO REAJUSTADO": "ORÇAMENTO REAJUSTADO (R$)",
        "ORCAMENTO REAJUSTADO": "ORÇAMENTO REAJUSTADO (R$)",
        "DESEMBOLSO ACUMULADO": "DESEMBOLSO ACUMULADO (R$)",
        "A PAGAR": "A PAGAR (R$)",
        "SALDO A INCORRER": "SALDO A INCORRER (R$)",
        "CUSTO FINAL": "CUSTO FINAL (R$)",
        "VARIAÇÃO": "VARIAÇÃO (R$)",
        "VARIACAO": "VARIAÇÃO (R$)",
    }

    out: dict[str, float | None] = {}
    r = title_row + 1
    while r <= ws.max_row:
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if is_blank(k):
            break
        key_norm = norm(k)
        key = mapping.get(key_norm, str(k).strip())
        out[key] = to_float(v)
        r += 1

    return out


def read_indice(ws: Worksheet) -> pd.DataFrame:
    hr, row_vals = _find_header_row(ws, ["MÊS", "ÍNDICE PROJETADO"], max_row=300, max_col=30)
    if hr is None or row_vals is None:
        return pd.DataFrame(columns=["MÊS", "ÍNDICE PROJETADO"])

    c_mes = _col_idx(row_vals, "MÊS")
    c_idx = _col_idx(row_vals, "ÍNDICE PROJETADO")
    if c_mes is None or c_idx is None:
        return pd.DataFrame(columns=["MÊS", "ÍNDICE PROJETADO"])

    data = []
    r = hr + 1
    while r <= ws.max_row:
        mes = ws.cell(r, c_mes).value
        if is_blank(mes):
            break
        data.append({"MÊS": to_month(mes), "ÍNDICE PROJETADO": to_float(ws.cell(r, c_idx).value)})
        r += 1

    df = pd.DataFrame(data).dropna(subset=["MÊS"])
    return df


def read_financeiro(ws: Worksheet) -> pd.DataFrame:
    hr, row_vals = _find_header_row(ws, ["MÊS", "DESEMBOLSO DO MÊS (R$)", "MEDIDO NO MÊS (R$)"], max_row=350, max_col=50)
    if hr is None or row_vals is None:
        return pd.DataFrame(columns=["MÊS", "DESEMBOLSO DO MÊS (R$)", "MEDIDO NO MÊS (R$)"])

    c_mes = _col_idx(row_vals, "MÊS")
    c_des = _col_idx(row_vals, "DESEMBOLSO DO MÊS (R$)")
    c_med = _col_idx(row_vals, "MEDIDO NO MÊS (R$)")
    if c_mes is None or c_des is None or c_med is None:
        return pd.DataFrame(columns=["MÊS", "DESEMBOLSO DO MÊS (R$)", "MEDIDO NO MÊS (R$)"])

    data = []
    r = hr + 1
    while r <= ws.max_row:
        mes = ws.cell(r, c_mes).value
        if is_blank(mes):
            break
        data.append(
            {
                "MÊS": to_month(mes),
                "DESEMBOLSO DO MÊS (R$)": to_float(ws.cell(r, c_des).value),
                "MEDIDO NO MÊS (R$)": to_float(ws.cell(r, c_med).value),
            }
        )
        r += 1

    df = pd.DataFrame(data).dropna(subset=["MÊS"])
    return df


def read_prazo(ws: Worksheet) -> pd.DataFrame:
    # Aceita “PREVISTO MENSAL (%)” novo + opcional “PLANEJADO ACUM. (%)”
    # Vamos achar a linha por MÊS + PLANEJADO MÊS + REALIZADO Mês (mínimo)
    hr, row_vals = _find_header_row(ws, ["MÊS", "PLANEJADO MÊS (%)", "REALIZADO Mês (%)"], max_row=600, max_col=40)
    if hr is None or row_vals is None:
        return pd.DataFrame(columns=["MÊS", "PLANEJADO ACUM. (%)", "PLANEJADO MÊS (%)", "PREVISTO MENSAL (%)", "REALIZADO Mês (%)"])

    c_mes = _col_idx(row_vals, "MÊS")
    c_plan_m = _col_idx(row_vals, "PLANEJADO MÊS (%)")
    c_real_m = _col_idx(row_vals, "REALIZADO Mês (%)")

    # opcionais
    c_plan_ac = _col_idx(row_vals, "PLANEJADO ACUM. (%)") or _col_idx(row_vals, "PLANEJADO ACUMULADO (%)")
    c_prev_m = _col_idx(row_vals, "PREVISTO MENSAL (%)") or _col_idx(row_vals, "PREVISTO MÊS (%)") or _col_idx(row_vals, "PREVISTO MES (%)")

    if c_mes is None or c_plan_m is None or c_real_m is None:
        return pd.DataFrame(columns=["MÊS", "PLANEJADO ACUM. (%)", "PLANEJADO MÊS (%)", "PREVISTO MENSAL (%)", "REALIZADO Mês (%)"])

    data = []
    r = hr + 1
    while r <= ws.max_row:
        mes = ws.cell(r, c_mes).value
        if is_blank(mes):
            break

        row = {
            "MÊS": to_month(mes),
            "PLANEJADO MÊS (%)": to_float(ws.cell(r, c_plan_m).value),
            "REALIZADO Mês (%)": to_float(ws.cell(r, c_real_m).value),
        }
        if c_plan_ac is not None:
            row["PLANEJADO ACUM. (%)"] = to_float(ws.cell(r, c_plan_ac).value)
        if c_prev_m is not None:
            row["PREVISTO MENSAL (%)"] = to_float(ws.cell(r, c_prev_m).value)

        data.append(row)
        r += 1

    df = pd.DataFrame(data).dropna(subset=["MÊS"])
    return df


def read_acrescimos_economias(ws: Worksheet) -> tuple[pd.DataFrame, pd.DataFrame]:
    # encontra a linha onde aparece “ACRÉSCIMOS” e “ECONOMIAS”
    base_row = None
    for r in range(1, min(ws.max_row, 800) + 1):
        a = norm(ws.cell(r, 1).value)
        g = norm(ws.cell(r, 7).value)
        if ("ACRÉSCIM" in a or "ACRESCIM" in a) and ("ECONOM" in g):
            base_row = r
            break

    # fallback: header DESCRIÇÃO dos dois lados
    if base_row is None:
        for r in range(1, min(ws.max_row, 800) + 1):
            if norm(ws.cell(r, 1).value) == "DESCRIÇÃO" and norm(ws.cell(r, 7).value) == "DESCRIÇÃO":
                base_row = r - 1
                break

    empty = pd.DataFrame(columns=["DESCRIÇÃO", "ORÇAMENTO INICIAL", "ORÇAMENTO REAJUSTADO", "CUSTO FINAL", "VARIAÇÃO", "JUSTIFICATIVAS"])
    if base_row is None:
        return empty, empty

    header_row = base_row + 1

    def map_cols(start: int, end: int) -> dict[str, int]:
        targets = {
            "DESCRIÇÃO": ["DESCRIÇÃO", "DESCRICAO"],
            "ORÇAMENTO INICIAL": ["ORÇAMENTO INICIAL", "ORCAMENTO INICIAL"],
            "ORÇAMENTO REAJUSTADO": ["ORÇAMENTO REAJUSTADO", "ORCAMENTO REAJUSTADO"],
            "CUSTO FINAL": ["CUSTO FINAL"],
            "VARIAÇÃO": ["VARIAÇÃO", "VARIACAO"],
            "JUSTIFICATIVAS": ["JUSTIFICATIVAS", "JUSTIFICATIVA"],
        }
        found: dict[str, int] = {}
        for c in range(start, end + 1):
            h = norm(ws.cell(header_row, c).value)
            for key, opts in targets.items():
                if key in found:
                    continue
                if h in [norm(x) for x in opts]:
                    found[key] = c
        return found

    left_cols = map_cols(1, 8)   # A..H (robusto)
    right_cols = map_cols(7, 14) # G..N (robusto)

    def read_side(colmap: dict[str, int]) -> pd.DataFrame:
        if "DESCRIÇÃO" not in colmap or "VARIAÇÃO" not in colmap:
            return empty.copy()

        data = []
        r = header_row + 1
        while r <= ws.max_row:
            desc = ws.cell(r, colmap["DESCRIÇÃO"]).value
            if is_blank(desc):
                break

            row = {
                "DESCRIÇÃO": str(desc).strip(),
                "ORÇAMENTO INICIAL": to_float(ws.cell(r, colmap.get("ORÇAMENTO INICIAL", -1)).value) if "ORÇAMENTO INICIAL" in colmap else None,
                "ORÇAMENTO REAJUSTADO": to_float(ws.cell(r, colmap.get("ORÇAMENTO REAJUSTADO", -1)).value) if "ORÇAMENTO REAJUSTADO" in colmap else None,
                "CUSTO FINAL": to_float(ws.cell(r, colmap.get("CUSTO FINAL", -1)).value) if "CUSTO FINAL" in colmap else None,
                "VARIAÇÃO": to_float(ws.cell(r, colmap["VARIAÇÃO"]).value),
                "JUSTIFICATIVAS": (str(ws.cell(r, colmap["JUSTIFICATIVAS"]).value).strip() if "JUSTIFICATIVAS" in colmap and not is_blank(ws.cell(r, colmap["JUSTIFICATIVAS"]).value) else ""),
            }
            data.append(row)
            r += 1

        df = pd.DataFrame(data)
        if df.empty:
            return empty.copy()
        return df

    return read_side(left_cols), read_side(right_cols)
