import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
import os
import openpyxl

# -------------------- Configura√ß√£o da p√°gina --------------------
st.set_page_config(
    page_title="ONE PAGE - ENGENHARIA",
    page_icon="üèóÔ∏è",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -------------------- Estilos CSS --------------------
st.markdown("""
<style>
.main-header {
    font-size: 2.8rem;
    color: #1E3A8A;
    font-weight: 800;
    margin-bottom: 2rem;
    padding-bottom: 0.5rem;
    border-bottom: 3px solid #3B82F6;
    text-align: center;
}
.sub-header {
    font-size: 1.8rem;
    color: #E5E8DD; /* Branco Nuvem */
    font-weight: 700;
    margin: 2rem 0 1.5rem 0;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid #5DAAAB; /* Azul C√©u */
}

.metric-card {
    background-color: #1d2f57; /* Azul Fundo do Mar */
    border-radius: 0.75rem;
    padding: 1rem;
    box-shadow: 0 4px 6px rgba(0,0,0,0.3);
    height: 100%;
    border-left: 5px solid #5DAAAB; /* Azul C√©u */
    margin-bottom: 1rem;
}
.metric-title {
    font-size: 1rem;
    color: #5DAAAB; /* Azul C√©u */
    font-weight: 600;
    margin-bottom: 0.5rem;
}
.metric-value {
    font-size: 1.6rem;
    color: #E5E8DD; /* Branco Nuvem */
    font-weight: 800;
}
.section-container {
    background-color: #1A253C; /* Azul Escuro */
    border-radius: 1rem;
    padding: 1.5rem;
    margin-bottom: 2rem;
    box-shadow: 0 4px 6px rgba(0,0,0,0.4);
}
.progress-wrapper {
    background-color: #1d2f57; /* Azul Escuro */
    border-radius: 20px;
    padding: 5px;
    width: 100%;
}
.progress-bar {
    height: 30px;
    border-radius: 20px;
    text-align: center;
    font-weight: bold;
    color: #E5E8DD; /* Branco Nuvem */
    line-height: 30px;
}
</style>
""", unsafe_allow_html=True)

# -------------------- Sidebar (filtro + logo empresa) --------------------
logo_empresa_path = "empresa_logo.png"
if os.path.exists(logo_empresa_path):
    st.sidebar.image(logo_empresa_path, width=200)

st.sidebar.markdown("### üìÇ Selecione a Obra - Jul/25")

file_path = "ONE_PAGE.xlsx"
if not os.path.exists(file_path):
    st.error("‚ùå Arquivo ONE_PAGE.xlsx n√£o encontrado no diret√≥rio!")
    st.stop()

excel_file = pd.ExcelFile(file_path)
sheet_names = excel_file.sheet_names
selected_sheet = st.sidebar.selectbox("Obra:", sheet_names)



# -------------------- Carregar dados --------------------
df = pd.read_excel(file_path, sheet_name=selected_sheet)
df_clean = df.iloc[:, [0, 1]].dropna()
df_clean.columns = ['Metrica', 'Valor']

dados = {str(row['Metrica']).strip(): row['Valor'] for _, row in df_clean.iterrows()}

def get_value(key, default="N/A"):
    return dados.get(key, default)

def format_money(value):
    if isinstance(value, (int, float)):
        return f"R$ {value:,.0f}".replace(',', '.')
    return str(value)

def format_percent(value):
    try:
        if isinstance(value, (int, float)):
            # Todos os valores do Excel em formato % chegam como decimais (0.85 = 85%)
            return f"{value*100:.2f}%"
        return str(value)
    except:
        return "N/A"

def to_float(val):
    if isinstance(val, str):
        try:
            return float(val.replace('R$', '').replace('.', '').replace(',', '.'))
        except:
            return 0
    return val if isinstance(val, (int,float)) else 0

# -------------------- Logo da obra --------------------
obra_logo_path = f"{selected_sheet}.png"
if os.path.exists(obra_logo_path):
    st.image(obra_logo_path, width=350)

# -------------------- M√©tricas Principais --------------------
st.markdown('<p class="sub-header">üìä Dados do Empreendimento</p>', unsafe_allow_html=True)
cols = st.columns(4)

cols[0].markdown(f'<div class="metric-card"><p class="metric-title">√Årea Constru√≠da (m¬≤)</p><p class="metric-value">{get_value("√Årea Constru√≠da (m¬≤)")}</p></div>', unsafe_allow_html=True)
cols[1].markdown(f'<div class="metric-card"><p class="metric-title">√Årea Privativa (m¬≤)</p><p class="metric-value">{get_value("√Årea Privativa (m¬≤)")}</p></div>', unsafe_allow_html=True)
cols[2].markdown(f'<div class="metric-card"><p class="metric-title">Efici√™ncia</p><p class="metric-value">{format_percent(get_value("Efici√™ncia"))}</p></div>',unsafe_allow_html=True)
cols[3].markdown(f'<div class="metric-card"><p class="metric-title">Unidades</p><p class="metric-value">{get_value("Unidades")}</p></div>', unsafe_allow_html=True)


# -------------------- Segunda linha de m√©tricas --------------------
cols2 = st.columns(4)

cols2[0].markdown(f'<div class="metric-card"><p class="metric-title">Rentabilidade Viabilidade</p><p class="metric-value">{format_percent(get_value("Rentabilidade Viabilidade"))}</p></div>', unsafe_allow_html=True)
cols2[1].markdown(f'<div class="metric-card"><p class="metric-title">Rentabilidade Projetada</p><p class="metric-value">{format_percent(get_value("Rentabilidade Projetada"))}</p></div>', unsafe_allow_html=True)
cols2[2].markdown(f'<div class="metric-card"><p class="metric-title">Custo √Årea Constru√≠da</p><p class="metric-value">{format_money(get_value("Custo √Årea Constru√≠da"))}</p></div>', unsafe_allow_html=True)
cols2[3].markdown(f'<div class="metric-card"><p class="metric-title">Custo √Årea Privativa</p><p class="metric-value">{format_money(get_value("Custo √Årea Privativa"))}</p></div>', unsafe_allow_html=True)


# -------------------- An√°lise Financeira --------------------
st.markdown('<p class="sub-header">üí∞ An√°lise Financeira</p>', unsafe_allow_html=True)

# Primeira linha
cols1 = st.columns(3)
cols1[0].markdown(f'<div class="metric-card"><p class="metric-title">Or√ßamento Base</p><p class="metric-value">{format_money(get_value("Or√ßamento Base"))}</p></div>', unsafe_allow_html=True)
cols1[1].markdown(f'<div class="metric-card"><p class="metric-title">Or√ßamento Reajustado</p><p class="metric-value">{format_money(get_value("Or√ßamento Reajustado"))}</p></div>', unsafe_allow_html=True)
cols1[2].markdown(f'<div class="metric-card"><p class="metric-title">Custo Final</p><p class="metric-value">{format_money(get_value("Custo Final"))}</p></div>', unsafe_allow_html=True)

# Segunda linha
cols2 = st.columns(4)
cols2[0].markdown(f'<div class="metric-card"><p class="metric-title">Desvio</p><p class="metric-value">{format_money(get_value("Desvio"))}</p></div>', unsafe_allow_html=True)
cols2[1].markdown(f'<div class="metric-card"><p class="metric-title">Desembolso</p><p class="metric-value">{format_money(get_value("Desembolso"))}</p></div>', unsafe_allow_html=True)
cols2[2].markdown(f'<div class="metric-card"><p class="metric-title">Saldo</p><p class="metric-value">{format_money(get_value("Saldo"))}</p></div>', unsafe_allow_html=True)
cols2[3].markdown(f'<div class="metric-card"><p class="metric-title">√çndice Econ√¥mico</p><p class="metric-value">{get_value("√çndice Econ√¥mico")}</p></div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# -------------------- Barra de progresso (Avan√ßo F√≠sico) --------------------
st.markdown('<p class="sub-header">üìÖ Avan√ßo F√≠sico</p>', unsafe_allow_html=True)

av_real_num = get_value("Avan√ßo F√≠sico Real", 0)
av_plan_num = get_value("Avan√ßo F√≠sico Planejado", 0)
aderencia_num = get_value("Ader√™ncia F√≠sica", 0)

st.markdown(f"""
<div class="progress-wrapper">
    <div class="progress-bar" style="width: {av_real_num*100:.0f}%; background: #3B82F6;">
        Real: {format_percent(av_real_num)}
    </div>
</div>
<p style="color:#EF4444;font-weight:600;">Planejado: {format_percent(av_plan_num)}</p>
<p style="color:#10B981;font-weight:600;">Ader√™ncia: {format_percent(aderencia_num)}</p>
""", unsafe_allow_html=True)


# -------------------- Linha do Tempo --------------------
st.markdown('<p class="sub-header">‚è∞ Linha do Tempo</p>', unsafe_allow_html=True)

# Dicion√°rio para meses em portugu√™s
meses_pt = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
    5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
    9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}

# Fun√ß√£o para formatar apenas m√™s/ano em portugu√™s
def format_month_year_pt(date_val):
    try:
        dt = pd.to_datetime(date_val)
        return f"{meses_pt[dt.month]}/{dt.year}"  # Ex: Jun/2025 ‚Üí Jun/2025
    except:
        return None

# Pegar datas
inicio = get_value("In√≠cio", None)
tend = get_value("Tend√™ncia", None)
prazo_concl = get_value("Prazo Conclus√£o", None)
prazo_cliente = get_value("Prazo Cliente", None)

# Criar cards para cada data
cards = [
    {"label": "In√≠cio", "date": format_month_year_pt(inicio), "color": "#3B82F6", "raw": inicio},
    {"label": "Tend√™ncia", "date": format_month_year_pt(tend), "color": "#F59E0B", "raw": tend},
    {"label": "Prazo Conclus√£o", "date": format_month_year_pt(prazo_concl), "color": "#10B981", "raw": prazo_concl},
    {"label": "Prazo Cliente", "date": format_month_year_pt(prazo_cliente), "color": "#EF4444", "raw": prazo_cliente}
]

# Mostrar cards coloridos
cols = st.columns(len(cards))
for col, card in zip(cols, cards):
    col.markdown(f"""
        <div style="background-color:{card['color']}; padding: 15px; border-radius: 10px; text-align:center; color:white;">
            <p style="margin:0; font-weight:bold;">{card['label']}</p>
            <p style="margin:0;">{card['date'] if card['date'] else 'N/A'}</p>
        </div>
    """, unsafe_allow_html=True)

# -------------------- Linha Temporal --------------------
valid_cards = [c for c in cards if c['raw'] is not None]
if len(valid_cards) >= 2:
    dates = [pd.to_datetime(c['raw']) for c in valid_cards]
    labels = [c['label'] for c in valid_cards]
    colors = [c['color'] for c in valid_cards]

    fig_timeline = go.Figure()
    # Linha base
    fig_timeline.add_trace(go.Scatter(
        x=[min(dates), max(dates)],
        y=[0, 0],
        mode='lines',
        line=dict(color='gray', width=3),
        showlegend=False
    ))

    # Pontos
    for date, label, color in zip(dates, labels, colors):
        fig_timeline.add_trace(go.Scatter(
            x=[date],
            y=[0],
            mode='markers+text',
            marker=dict(size=15, color=color),
            text=[label],
            textposition='top center',
            name=label,
            textfont=dict(color='white', size=12)
        ))

    fig_timeline.update_layout(
        title='Cronograma da Obra',
        showlegend=False,
        height=200,
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(showgrid=False, zeroline=False, title=''),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        font=dict(color='white')
    )
    st.plotly_chart(fig_timeline, use_container_width=True)
else:
    st.info("N√£o h√° datas suficientes para criar a linha do tempo.")


# -------------------- Status do Andamento da Obra --------------------
st.markdown('<p class="sub-header">üìù Status Andamento da Obra</p>', unsafe_allow_html=True)

# Filtrar status existentes
status_rows = df_clean[df_clean['Metrica'].str.strip() == "Status Andamento Obra"]

with st.expander("üìå Ver / Editar Status", expanded=False):
    # Mostrar status existentes
    if not status_rows.empty:
        for i, status in enumerate(status_rows['Valor'], 1):
            st.markdown(f"**{i}.** {status}")

    st.markdown("---")
    # Input para adicionar novo status
    new_status = st.text_area("Adicionar novo status", placeholder="Digite aqui o novo status...")

    if st.button("üíæ Salvar Status"):
        if new_status.strip() != "":
            import openpyxl

            file_path = "ONE_PAGE.xlsx"
            wb = openpyxl.load_workbook(file_path)
            ws = wb[selected_sheet]

            # Procurar √∫ltima linha ocupada
            last_row = ws.max_row
            while last_row > 0 and ws.cell(row=last_row, column=1).value is None:
                last_row -= 1
            next_row = last_row + 1

            # Inserir na coluna A (t√≠tulo) e B (conte√∫do)
            ws.cell(row=next_row, column=1, value="Status Andamento Obra")
            ws.cell(row=next_row, column=2, value=new_status.strip())

            # Salvar de volta no Excel
            wb.save(file_path)

            # Recarregar dataframe para mostrar imediatamente
            df_updated = pd.read_excel(file_path, sheet_name=selected_sheet)
            df_clean = df_updated.iloc[:, [0, 1]].dropna()
            df_clean.columns = ['Metrica', 'Valor']
            status_rows = df_clean[df_clean['Metrica'].str.strip() == "Status Andamento Obra"]

            st.success("‚úÖ Novo status salvo com sucesso!")

            # Mostrar lista atualizada
            for i, status in enumerate(status_rows['Valor'], 1):
                st.markdown(f"**{i}.** {status}")
        else:
            st.warning("‚ö†Ô∏è Digite algum valor antes de salvar.")



st.markdown("""---""")

st.markdown(
    """
    <div style='text-align: center; font-size: 14px; color: gray; padding-top: 20px;'>
        <i>"Inspirados pelo que te faz bem"</i>
        <br>
        Desenvolvido por <b>Matheus Vinicio</b> ‚Äî Engenharia
        <br>
        ¬© 2025 <a href='https://www.rioave.com.br/' target='_blank' style='color: gray; text-decoration: none;'><b>RIO AVE</b></a>
    </div>
    """,
    unsafe_allow_html=True
)

