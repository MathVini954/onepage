from __future__ import annotations

from pathlib import Path
from typing import Any

import re
import unicodedata

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.datetime import from_excel


# ============================================================
# Normalização / parsing
# ============================================================
_PT_MONTH = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))


def _norm(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip().replace("\u00a0", " ")
    s = " ".join(s.split())
    s = _strip_accents(s).upper()
    return s


def _is_blank(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, str) and not x.strip():
        return True
    return False


def _to_float(x: Any) -> float | None:
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        s = str(x).strip()
        if not s:
            return None

        s = s.replace("R$", "").replace(" ", "")

        # caso "1,0055" (vírgula decimal sem milhar)
        if "," in s and "." not in s:
            try:
                return float(s.replace(",", "."))
            except Exception:
                return None

        # caso BR "1.234.567,89"
        try:
            s2 = s.replace(".", "").replace(",", ".")
            return float(s2)
        except Exception:
            return None


def _to_month(x: Any) -> pd.Timestamp | None:
    """
    Converte:
      - datetime/date
      - serial do Excel (ex: 45292)
      - strings: 'jan/2026', 'fev/2026', 'Jan.26', 'JAN/26', '01/01/2026'
    Retorna Timestamp (1º dia do mês) ou None.
    """
    try:
        if x is None:
            return None

        # datetime/date
        if hasattr(x, "year") and hasattr(x, "month"):
            dt = pd.Timestamp(x)
            if pd.isna(dt):
                return None
            return pd.Timestamp(year=dt.year, month=dt.month, day=1)

        # serial do Excel
        if isinstance(x, (int, float)):
            if x <= 0:
                return None
            dt = from_excel(x)
            dt = pd.Timestamp(dt)
            return pd.Timestamp(year=dt.year, month=dt.month, day=1)

        # string
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return None

            up = _norm(s)

            # "JAN/2026" "FEV/26" "JAN.26" "SET-2025"
            m = re.match(r"^([A-Z]{3})[./\- ]*(\d{2,4})$", up)
            if m:
                mon = _PT_MONTH.get(m.group(1))
                yy = int(m.group(2))
                if mon:
                    year = yy if yy >= 100 else (2000 + yy)
                    return pd.Timestamp(year=year, month=mon, day=1)

            # tenta parse padrão
            dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
            if pd.notna(dt):
                dt = pd.Timestamp(dt)
                return pd.Timestamp(year=dt.year, month=dt.month, day=1)

            return None

        # fallback geral
        dt = pd.to_datetime(x, errors="coerce")
        if pd.notna(dt):
            dt = pd.Timestamp(dt)
            return pd.Timestamp(year=dt.year, month=dt.month, day=1)

        return None
    except Exception:
        return None


# ============================================================
# Workbook helpers
# ============================================================
def load_wb(path: str | Path):
    path = Path(path)
    keep = path.suffix.lower() == ".xlsm"
    return openpyxl.load_workbook(path, data_only=True, keep_vba=keep)


def sheetnames(wb) -> list[str]:
    """
    Retorna apenas as abas das OBRAS.
    Ignora LEIA-ME e ORÇAMENTO_RESUMO (que é só para a tela Resumo Obras).
    """
    IGNORAR = {"LEIA-ME", "LEIA ME", "README", "ORÇAMENTO_RESUMO", "ORCAMENTO_RESUMO"}
    out: list[str] = []
    for name in wb.sheetnames:
        if _norm(name) in {_norm(x) for x in IGNORAR}:
            continue
        if _norm(name).startswith("_"):
            continue
        out.append(name)
    return out


# ============================================================
# Leitores (mesma lógica de antes: por aba da OBRA)
# ============================================================
def read_resumo_financeiro(ws: Worksheet) -> dict[str, float | None]:
    wanted = {
        "ORÇAMENTO INICIAL (R$)",
        "ORÇAMENTO REAJUSTADO (R$)",
        "DESEMBOLSO ACUMULADO (R$)",
        "A PAGAR (R$)",
        "SALDO A INCORRER (R$)",
        "CUSTO FINAL (R$)",
        "VARIAÇÃO (R$)",
    }
    out: dict[str, float | None] = {}
    for r in range(1, min(ws.max_row or 1, 200) + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue
        lab = str(label).strip()
        if lab in wanted:
            out[lab] = _to_float(ws.cell(r, 2).value)
    return out


def read_indice(ws: Worksheet) -> pd.DataFrame:
    """
    Lê a tabela 'MÊS' x 'ÍNDICE PROJETADO' ignorando o título mesclado.
    Procura a linha que contém MES e INDICE PROJETADO na mesma linha.
    """
    max_r = min(ws.max_row or 1, 400)
    max_c = min(ws.max_column or 1, 160)

    header_row = None
    col_mes = None
    col_idx = None

    for r in range(1, max_r + 1):
        c_mes_tmp = None
        c_idx_tmp = None
        for c in range(1, max_c + 1):
            v = _norm(ws.cell(r, c).value)
            if v == "MES":
                c_mes_tmp = c
            if ("INDICE" in v) and ("PROJETADO" in v):
                c_idx_tmp = c
        if c_mes_tmp is not None and c_idx_tmp is not None:
            header_row = r
            col_mes = c_mes_tmp
            col_idx = c_idx_tmp
            break

    if header_row is None or col_mes is None or col_idx is None:
        return pd.DataFrame(columns=["MÊS", "ÍNDICE PROJETADO"])

    rows = []
    blank_mes_run = 0

    for r in range(header_row + 1, (ws.max_row or header_row + 1) + 1):
        mes = ws.cell(r, col_mes).value
        idx = ws.cell(r, col_idx).value

        if _is_blank(mes):
            blank_mes_run += 1
            if blank_mes_run >= 4:
                break
            continue
        blank_mes_run = 0

        m = _to_month(mes)
        v = _to_float(idx)

        # ignora vazios/zeros
        if m is None or v is None or float(v) == 0.0:
            continue

        rows.append((m, v))

    df = pd.DataFrame(rows, columns=["MÊS", "ÍNDICE PROJETADO"])
    if not df.empty:
        df = df.sort_values("MÊS")
    return df


def read_financeiro(ws: Worksheet) -> pd.DataFrame:
    """
    Procura header: 'DESEMBOLSO DO MÊS' e 'MEDIDO NO MÊS' e lê a tabela.
    """
    max_r = min(ws.max_row or 1, 600)
    max_c = min(ws.max_column or 1, 160)

    header_row = None
    col_mes = col_des = col_med = None

    for r in range(1, max_r + 1):
        found_des = found_med = False
        c_mes_tmp = None
        c_des_tmp = None
        c_med_tmp = None

        for c in range(1, max_c + 1):
            v = _norm(ws.cell(r, c).value)
            if v == "MES":
                c_mes_tmp = c
            if "DESEMBOLSO" in v and "MES" in v:
                found_des = True
                c_des_tmp = c
            if "MEDIDO" in v and "MES" in v:
                found_med = True
                c_med_tmp = c

        if found_des and found_med and c_mes_tmp is not None:
            header_row = r
            col_mes = c_mes_tmp
            col_des = c_des_tmp
            col_med = c_med_tmp
            break

    if header_row is None or col_mes is None or col_des is None or col_med is None:
        return pd.DataFrame(columns=["MÊS", "DESEMBOLSO DO MÊS (R$)", "MEDIDO NO MÊS (R$)"])

    rows = []
    blank_mes_run = 0
    for r in range(header_row + 1, (ws.max_row or header_row + 1) + 1):
        mes = ws.cell(r, col_mes).value
        des = ws.cell(r, col_des).value
        med = ws.cell(r, col_med).value

        if _is_blank(mes):
            blank_mes_run += 1
            if blank_mes_run >= 4:
                break
            continue
        blank_mes_run = 0

        m = _to_month(mes)
        if m is None:
            continue

        rows.append((m, _to_float(des), _to_float(med)))

    df = pd.DataFrame(rows, columns=["MÊS", "DESEMBOLSO DO MÊS (R$)", "MEDIDO NO MÊS (R$)"])
    if not df.empty:
        df = df.sort_values("MÊS")
    return df


def read_prazo(ws: Worksheet) -> pd.DataFrame:
    """
    Procura header do prazo por:
      MES, PLANEJADO (MES/ACUM), REALIZADO e PREVISTO MENSAL (opcional)
    Aceita header como "PREVISTO MENSAL(%)" também.
    """
    max_r = min(ws.max_row or 1, 900)
    max_c = min(ws.max_column or 1, 180)

    header_row = None
    c_mes = c_pa = c_pm = c_rm = c_prev = None

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = _norm(ws.cell(r, c).value)

            if v == "MES":
                c_mes = c

            if "PLANEJADO" in v and "ACUM" in v:
                c_pa = c

            # Planejado mês: precisa ter PLANEJADO + MES e NÃO ser ACUM
            if "PLANEJADO" in v and "MES" in v and "ACUM" not in v:
                c_pm = c

            if "REALIZADO" in v:
                c_rm = c

            if "PREVISTO" in v and "MENSAL" in v:
                c_prev = c

        if c_mes is not None and c_pm is not None and c_rm is not None:
            header_row = r
            break

    if header_row is None or c_mes is None:
        return pd.DataFrame(
            columns=["MÊS", "PLANEJADO ACUM. (%)", "PLANEJADO MÊS (%)", "REALIZADO Mês (%)", "PREVISTO MENSAL (%)"]
        )

    rows = []
    blank_mes_run = 0
    for r in range(header_row + 1, (ws.max_row or header_row + 1) + 1):
        mes = ws.cell(r, c_mes).value

        if _is_blank(mes):
            blank_mes_run += 1
            if blank_mes_run >= 6:
                break
            continue
        blank_mes_run = 0

        m = _to_month(mes)
        if m is None:
            continue

        pa = ws.cell(r, c_pa).value if c_pa else None
        pm = ws.cell(r, c_pm).value if c_pm else None
        rm = ws.cell(r, c_rm).value if c_rm else None
        pv = ws.cell(r, c_prev).value if c_prev else None

        rows.append((m, _to_float(pa), _to_float(pm), _to_float(rm), _to_float(pv)))

    df = pd.DataFrame(
        rows,
        columns=["MÊS", "PLANEJADO ACUM. (%)", "PLANEJADO MÊS (%)", "REALIZADO Mês (%)", "PREVISTO MENSAL (%)"],
    )
    if not df.empty:
        df = df.sort_values("MÊS")
    return df


def read_acrescimos_economias(ws: Worksheet) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Acha uma linha que tenha 2x 'DESCRIÇÃO' (acr/eco).
    Lê 6 colunas por lado:
      DESCRIÇÃO, ORÇAMENTO INICIAL, ORÇAMENTO REAJUSTADO, CUSTO FINAL, VARIAÇÃO, JUSTIFICATIVAS
    """
    max_r = min(ws.max_row or 1, 2000)
    max_c = min(ws.max_column or 1, 220)

    header_row = None
    start1 = start2 = None

    for r in range(1, max_r + 1):
        desc_cols = []
        for c in range(1, max_c + 1):
            v = _norm(ws.cell(r, c).value)
            if v == "DESCRICAO" or "DESCRICAO" in v:
                desc_cols.append(c)
        if len(desc_cols) >= 2:
            header_row = r
            start1, start2 = desc_cols[0], desc_cols[1]
            break

    cols = ["DESCRIÇÃO", "ORÇAMENTO INICIAL", "ORÇAMENTO REAJUSTADO", "CUSTO FINAL", "VARIAÇÃO", "JUSTIFICATIVAS"]
    if header_row is None or start1 is None or start2 is None:
        return pd.DataFrame(columns=cols), pd.DataFrame(columns=cols)

    def read_side(start_col: int) -> pd.DataFrame:
        rows = []
        blank_run = 0
        for r in range(header_row + 1, (ws.max_row or header_row + 1) + 1):
            vals = [ws.cell(r, start_col + i).value for i in range(6)]
            if all(_is_blank(v) for v in vals):
                blank_run += 1
                if blank_run >= 10:
                    break
                continue
            blank_run = 0

            desc = vals[0]
            if _is_blank(desc):
                continue

            rows.append(
                (
                    str(desc).strip(),
                    _to_float(vals[1]),
                    _to_float(vals[2]),
                    _to_float(vals[3]),
                    _to_float(vals[4]),
                    "" if vals[5] is None else str(vals[5]).strip(),
                )
            )

        return pd.DataFrame(rows, columns=cols)

    df_acres = read_side(start1)
    df_econ = read_side(start2)
    return df_acres, df_econ


# ============================================================
# NOVO: ORÇAMENTO_RESUMO (somente para a aba "Resumo Obras")
# ============================================================
def read_orcamento_resumo(wb) -> pd.DataFrame:
    """
    Lê a aba ORÇAMENTO_RESUMO (não interfere em nada do Dashboard/Justificativas).
    Esperado: primeira coluna OBRA, depois meses, e última coluna VARIAÇÃO.
    """
    candidates = ["ORÇAMENTO_RESUMO", "ORCAMENTO_RESUMO"]
    sheet = None
    for s in candidates:
        if s in wb.sheetnames:
            sheet = s
            break
    if sheet is None:
        return pd.DataFrame()

    ws = wb[sheet]

    # acha o header procurando "OBRA" em alguma célula da linha
    header_row = None
    max_r = min(ws.max_row or 1, 120)
    max_c = min(ws.max_column or 1, 250)

    for r in range(1, max_r + 1):
        found = False
        for c in range(1, max_c + 1):
            if _norm(ws.cell(r, c).value) == "OBRA":
                found = True
                break
        if found:
            header_row = r
            break

    if header_row is None:
        return pd.DataFrame()

    # lê cabeçalhos até acabar (colunas)
    headers = []
    last_col = 0
    for c in range(1, max_c + 1):
        v = ws.cell(header_row, c).value
        if v is None or str(v).strip() == "":
            if last_col > 0:
                # se já começou e encontrou vazio, encerra
                break
            continue
        headers.append(str(v).strip())
        last_col = c

    if not headers:
        return pd.DataFrame()

    # lê linhas até acabar
    data = []
    empty_streak = 0
    for r in range(header_row + 1, header_row + 1 + 5000):
        row = []
        all_empty = True
        for c in range(1, last_col + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                all_empty = False
            row.append(v)

        if all_empty:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue

        empty_streak = 0
        data.append(row)

    df = pd.DataFrame(data, columns=headers)

    if "OBRA" in df.columns:
        df["OBRA"] = df["OBRA"].astype(str).str.strip()

    return df
